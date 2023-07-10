import os

if input("Do you want to install library? (y/n) ").lower() == 'y':
    if input("Do you want to install `openpyxl`? (y/n) ").lower() == 'y':
        os.system("pip3 install openpyxl")

    if input("Do you want to install `clipboard`? (y/n) ").lower() == 'y':
        os.system("pip3 install clipboard")

import clipboard
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.simplefilter("ignore")

def is_table_name(table_name):
    except_list = ["Type", "ERDiagram"]
    if table_name is None:
        return None
    elif table_name in except_list:
        return None
    else:
        return table_name

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

table_set = set()
try:
    wb = load_workbook(input("input xlsx path: ")) # sample.xlsx 에서 wb를 불러옴
    ws = wb.active # 활성화된 Sheet

    col_D = ws["D"]
    for cell in col_D:
        table_name = is_table_name(cell.value)

        if table_name:
            table_name = table_name.removesuffix("_TB")
            table_set.add(table_name)

    output_text = ""
    for table_name in table_set:
        print(table_name)
        output_text += table_name + "\n"

    clipboard.copy(output_text)
    print(f"\n\033[94mtotal amount is {len(table_set)}\033[0m")
    print("\033[94mResult Copied!\033[0m")

except Exception as e:
    print(f"\033[91m{e}\033[0m")



while True:
    print("\n\n[Similar Checker]")
    input_table_name = input("table name: ")

    similar_dict = {}
    for table in table_set:
        similar_dict[table] = similar(input_table_name, table)

    sorted_dict = sorted(similar_dict.items(), key = lambda item: item[1], reverse = True)
    for data in sorted_dict:
        print(data)
    most_similar_word = sorted_dict[0][0]
    most_similarity = sorted_dict[0][1]

    if most_similarity == 0:
        print("\033[91mThere's not most similar word.\033[0m")
    else:
        print(f"\033[94mcopied most similar word : {most_similar_word} {round(most_similarity * 100)}%\033[0m")
